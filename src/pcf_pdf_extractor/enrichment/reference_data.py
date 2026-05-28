import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, field_validator

from pcf_pdf_extractor.domain import BooleanRequirementCheck, PCFRecord, PcfValueRequirementCheck, PcfValueResult


class JsonLlmClient(Protocol):
    """Minimal client interface needed by the reference-data enrichment step."""

    def complete_json(
        self,
        *,
        system_prompt: str,
        user_prompt: str,
        response_schema: dict[str, Any],
    ) -> dict[str, Any]:
        """Return a JSON object produced by an LLM."""


class BafuMappingResponse(BaseModel):
    """Strict JSON shape returned by the BAFU/Eclasses mapping prompt."""

    model_config = ConfigDict(extra="forbid")

    bafu_row: int | bool
    oil_gas_relevant: bool
    clarification_needed: bool
    reason: str

    @field_validator("bafu_row")
    @classmethod
    def bafu_row_cannot_be_true(cls, value: int | bool) -> int | bool:
        if value is True:
            raise ValueError("bafu_row must be an Excel row number or false")
        return value


@dataclass(frozen=True)
class BafuRow:
    """One usable row from the BAFU reference workbook."""

    row_number: int
    product: str
    unit: str
    gwp100_value: float
    gwp100_unit: str


class BafuReferenceData:
    """Read and query the BAFU workbook."""

    def __init__(self, rows: list[BafuRow]) -> None:
        self._rows = rows
        self._rows_by_number = {row.row_number: row for row in rows}

    @classmethod
    def from_excel(cls, path: Path) -> "BafuReferenceData":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        header_row = _find_header_row(sheet.iter_rows(values_only=True))
        header_cells = next(
            sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
        header_values = [_normalize_header(value) for value in header_cells]
        product_index = _header_index(header_values, "product")
        unit_index = _header_index(header_values, "unit")
        gwp_index = _header_index_containing(header_values, "gwp 100")
        gwp_unit = _gwp_unit_from_header(str(header_cells[gwp_index] or ""))

        rows: list[BafuRow] = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=False):
            product = row[product_index].value
            unit = row[unit_index].value
            gwp100_value = row[gwp_index].value
            if not isinstance(product, str) or not product.strip():
                continue
            if unit is None or gwp100_value is None:
                continue
            try:
                value = float(gwp100_value)
            except (TypeError, ValueError):
                continue

            unit_text = str(unit).strip()
            rows.append(
                BafuRow(
                    row_number=row[product_index].row,
                    product=product.strip(),
                    unit=unit_text,
                    gwp100_value=value,
                    gwp100_unit=f"{gwp_unit}/{unit_text}" if unit_text else gwp_unit,
                )
            )
        return cls(rows)

    def row_by_number(self, row_number: int) -> BafuRow | None:
        return self._rows_by_number.get(row_number)

    def candidate_rows(
        self,
        *,
        product_name: str,
        production_location: str | None = None,
        limit: int = 120,
    ) -> list[BafuRow]:
        query_variants = _query_variants(product_name)
        region_hint = _region_hint(production_location)
        scored_rows: list[tuple[float, BafuRow]] = []
        for row in self._rows:
            score = max(_score_candidate(query, row.product) for query in query_variants)
            row_region = _region_from_product(row.product)
            if region_hint and row_region == region_hint:
                score += 0.12
            if score > 0:
                scored_rows.append((score, row))

        scored_rows.sort(key=lambda item: (-item[0], item[1].row_number))
        return [row for _, row in scored_rows[:limit]]


class OilGasReferenceData:
    """Read the Oil & Gas relevance file."""

    def __init__(self, entries: list[str]) -> None:
        self.entries = entries

    @classmethod
    def from_text(cls, path: Path) -> "OilGasReferenceData":
        entries = [
            line.strip()
            for line in path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        return cls(entries)


class ReferenceDataEnricher:
    """Add BAFU expected GWP and Eclasses Oil & Gas relevance to a PCF record."""

    def __init__(
        self,
        *,
        client: JsonLlmClient,
        bafu_reference: BafuReferenceData,
        oil_gas_reference: OilGasReferenceData,
        mapping_prompt: str,
        candidate_limit: int = 120,
    ) -> None:
        self._client = client
        self._bafu_reference = bafu_reference
        self._oil_gas_reference = oil_gas_reference
        self._mapping_prompt = mapping_prompt
        self._candidate_limit = candidate_limit

    @classmethod
    def from_paths(
        cls,
        *,
        client: JsonLlmClient,
        bafu_path: Path,
        oil_gas_path: Path,
        prompt_path: Path,
        candidate_limit: int = 120,
    ) -> "ReferenceDataEnricher":
        return cls(
            client=client,
            bafu_reference=BafuReferenceData.from_excel(bafu_path),
            oil_gas_reference=OilGasReferenceData.from_text(oil_gas_path),
            mapping_prompt=prompt_path.read_text(encoding="utf-8"),
            candidate_limit=candidate_limit,
        )

    def enrich(self, record: PCFRecord) -> PCFRecord:
        if not record.product_name:
            record.expected_gwp100_value = PcfValueRequirementCheck(
                fulfilled=False,
                reason="No product_name was available for BAFU mapping.",
                evidence=None,
                result=None,
            )
            record.oil_gas_relevant = BooleanRequirementCheck(
                fulfilled=False,
                reason="No product_name was available, so Eclasses relevance was set to false.",
                evidence=None,
                result=False,
            )
            record.extraction_notes.append(
                "Reference data enrichment skipped because product_name is missing."
            )
            self._compute_checks(record)
            return record

        production_location = record.minimum_requirements.production_location.result
        candidates = self._bafu_reference.candidate_rows(
            product_name=record.product_name,
            production_location=production_location,
            limit=self._candidate_limit,
        )
        mapping_payload = self._client.complete_json(
            system_prompt=self._mapping_prompt,
            user_prompt=self._user_prompt(record, candidates),
            response_schema=BafuMappingResponse.model_json_schema(),
        )
        mapping = BafuMappingResponse.model_validate(mapping_payload)
        record.oil_gas_relevant = BooleanRequirementCheck(
            fulfilled=True,
            reason=(
                "Mapped from Eclasses relevance check in prompt_mapping."
                if mapping.oil_gas_relevant
                else "Not found in Eclasses relevance list by mapping step."
            ),
            evidence=None,
            result=mapping.oil_gas_relevant,
        )

        if mapping.clarification_needed:
            record.expected_gwp100_value = PcfValueRequirementCheck(
                fulfilled=False,
                reason=f"Clarification needed: {mapping.reason}",
                evidence=None,
                result=None,
            )
            record.extraction_notes.append(
                f"BAFU mapping requires clarification: {mapping.reason}"
            )
            self._compute_checks(record)
            return record

        if mapping.bafu_row is False:
            record.expected_gwp100_value = PcfValueRequirementCheck(
                fulfilled=False,
                reason="No safe BAFU row match was found.",
                evidence=None,
                result=None,
            )
            record.extraction_notes.append("No safe BAFU expected GWP 100 match was found.")
            self._compute_checks(record)
            return record

        bafu_row = self._bafu_reference.row_by_number(mapping.bafu_row)
        candidate_row_numbers = {candidate.row_number for candidate in candidates}
        if bafu_row is None or bafu_row.row_number not in candidate_row_numbers:
            record.expected_gwp100_value = PcfValueRequirementCheck(
                fulfilled=False,
                reason=f"Mapped row {mapping.bafu_row} is not in the candidate set.",
                evidence=None,
                result=None,
            )
            record.extraction_notes.append(
                f"BAFU mapping returned unavailable row {mapping.bafu_row}."
            )
            self._compute_checks(record)
            return record

        record.expected_gwp100_value = PcfValueRequirementCheck(
            fulfilled=True,
            reason=f"Mapped to BAFU row {bafu_row.row_number}: {bafu_row.product}.",
            evidence=None,
            result=PcfValueResult(
                value=bafu_row.gwp100_value,
                unit=bafu_row.gwp100_unit,
            ),
        )
        self._compute_checks(record)
        return record

    def _compute_checks(self, record: PCFRecord) -> None:
        actual = record.minimum_requirements.gwp100_excluding_biogenic.result
        expected = (
            record.expected_gwp100_value.result
            if record.expected_gwp100_value is not None
            else None
        )
        if actual is not None and expected is not None:
            if actual.value == 0:
                record.is_benchmarch_ok = expected.value == 0
            else:
                diff_ratio = abs(expected.value - actual.value) / abs(actual.value)
                record.is_benchmarch_ok = diff_ratio <= 0.30
        else:
            record.is_benchmarch_ok = False

        has_oil_gas_update = record.minimum_requirements.oil_and_gas_update.result is True
        has_secondary_databases = bool(record.minimum_requirements.secondary_databases.result)
        oil_gas_relevant = (
            record.oil_gas_relevant.result is True
            if record.oil_gas_relevant is not None
            else False
        )
        record.oil_and_gas_check_ok = (
            True
            if not oil_gas_relevant
            else has_oil_gas_update or has_secondary_databases
        )

    def _user_prompt(self, record: PCFRecord, candidates: list[BafuRow]) -> str:
        input_json = record.model_dump(mode="json")
        candidate_payload = [
            {
                "excel_row": candidate.row_number,
                "product": candidate.product,
                "unit": candidate.unit,
                "gwp100_value": candidate.gwp100_value,
                "gwp100_unit": candidate.gwp100_unit,
            }
            for candidate in candidates
        ]
        return (
            "Use the mapping instructions from the system prompt. The BAFU workbook has "
            "been pre-filtered by the application; select bafu_row only from "
            "BAFU_CANDIDATE_ROWS. If no candidate row is safe, return false or ask for "
            "clarification as instructed.\n\n"
            "INPUT_JSON:\n"
            f"{json.dumps(input_json, indent=2, ensure_ascii=False)}\n\n"
            "BAFU_CANDIDATE_ROWS:\n"
            f"{json.dumps(candidate_payload, indent=2, ensure_ascii=False)}\n\n"
            "OIL_GAS_RELEVANCE_ENTRIES:\n"
            f"{json.dumps(self._oil_gas_reference.entries, indent=2, ensure_ascii=False)}"
        )


def _find_header_row(rows: Any) -> int:
    for row_number, row in enumerate(rows, start=1):
        normalized = [_normalize_header(value) for value in row]
        if "product" in normalized and "unit" in normalized:
            return row_number
    raise ValueError("Could not find BAFU header row with Product and Unit columns")


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _header_index(headers: list[str], name: str) -> int:
    try:
        return headers.index(name)
    except ValueError as exc:
        raise ValueError(f"Could not find BAFU column {name!r}") from exc


def _header_index_containing(headers: list[str], name: str) -> int:
    for index, header in enumerate(headers):
        if name in header:
            return index
    raise ValueError(f"Could not find BAFU column containing {name!r}")


def _gwp_unit_from_header(header: str) -> str:
    match = re.search(r"\[([^\]]+)\]", header)
    if match:
        return match.group(1).strip()
    return "kg CO2 eq"


def _query_variants(product_name: str) -> list[str]:
    variants = {product_name, _strip_concentration(product_name)}
    normalized = _normalize_text(product_name)
    for key, aliases in _CHEMICAL_ALIASES.items():
        if key in normalized:
            variants.update(aliases)
    return [variant for variant in variants if variant]


def _strip_concentration(value: str) -> str:
    return re.sub(r"\b\d+(?:[.,]\d+)?\s*%\b", "", value).strip(" ,;-/")


def _score_candidate(query: str, product: str) -> float:
    normalized_query = _normalize_text(query)
    normalized_product = _normalize_text(_strip_region(product))
    if not normalized_query or not normalized_product:
        return 0

    score = 0.0
    if normalized_query == normalized_product:
        score += 1.0
    if normalized_query in normalized_product:
        score += 0.82

    query_tokens = _meaningful_tokens(normalized_query)
    product_tokens = _meaningful_tokens(normalized_product)
    if query_tokens and product_tokens:
        overlap = len(query_tokens & product_tokens)
        score += overlap / len(query_tokens | product_tokens)
        if query_tokens <= product_tokens:
            score += 0.45

    query_concentrations = _concentration_tokens(query)
    product_concentrations = _concentration_tokens(product)
    if query_concentrations:
        if query_concentrations & product_concentrations:
            score += 0.25
        else:
            score -= 0.2

    return score


def _meaningful_tokens(value: str) -> set[str]:
    stopwords = {
        "at",
        "by",
        "for",
        "from",
        "in",
        "of",
        "plant",
        "regional",
        "storehouse",
        "the",
        "to",
        "water",
        "with",
    }
    return {
        token
        for token in value.split()
        if len(token) > 1 and token not in stopwords
    }


def _concentration_tokens(value: str) -> set[str]:
    return {
        match.group(1).replace(",", ".")
        for match in re.finditer(r"\b(\d+(?:[.,]\d+)?)\s*%\b", value)
    }


def _region_hint(value: str | None) -> str | None:
    if not value:
        return None
    normalized = _normalize_text(value)
    tokens = set(normalized.split())
    region_codes = {
        "ch": "CH",
        "eu": "RER",
        "glo": "GLO",
        "rer": "RER",
        "row": "RoW",
        "us": "US",
        "usa": "US",
    }
    for key, region in region_codes.items():
        if key in tokens:
            return region

    known_regions = {
        "switzerland": "CH",
        "suisse": "CH",
        "france": "RER",
        "germany": "RER",
        "europe": "RER",
        "united states": "US",
        "north america": "RNA",
        "global": "GLO",
    }
    for key, region in known_regions.items():
        if key in normalized:
            return region
    return None


def _region_from_product(product: str) -> str | None:
    match = re.search(r"\{([^}]+)\}", product)
    return match.group(1).strip() if match else None


def _strip_region(product: str) -> str:
    return re.sub(r"\{[^}]+\}", "", product)


def _normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_value = ascii_value.replace("&", " and ")
    ascii_value = re.sub(r"[^a-zA-Z0-9%]+", " ", ascii_value)
    return re.sub(r"\s+", " ", ascii_value.lower()).strip()


_CHEMICAL_ALIASES = {
    "acetic acid": ["acetic acid", "essigsaure"],
    "ammonia": ["ammonia", "ammoniak"],
    "ammonium hydroxide": ["ammonium hydroxide", "ammoniumhydroxid", "ammonia solution"],
    "caustic soda": ["sodium hydroxide", "caustic soda", "natronlauge", "naoh"],
    "ethylene glycol": ["ethylene glycol", "ethylenglykol", "1 2 ethandiol", "meg"],
    "isopropanol": ["isopropanol", "isopropyl alcohol", "propan 2 ol"],
    "naoh": ["sodium hydroxide", "caustic soda", "natronlauge", "naoh"],
    "natronlauge": ["sodium hydroxide", "caustic soda", "natronlauge", "naoh"],
    "nitric acid": ["nitric acid", "salpetersaure"],
    "salpetersaure": ["nitric acid", "salpetersaure"],
    "schwefelsaure": ["sulfuric acid", "sulphuric acid", "schwefelsaure"],
    "sodium hydroxide": ["sodium hydroxide", "caustic soda", "natronlauge", "naoh"],
    "sulfuric acid": ["sulfuric acid", "sulphuric acid", "schwefelsaure"],
}
