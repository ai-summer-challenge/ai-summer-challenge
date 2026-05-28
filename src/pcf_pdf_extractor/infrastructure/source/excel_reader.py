from pathlib import Path
from typing import Any

from pcf_pdf_extractor.infrastructure.source.document import SourceDocument


class ExcelSourceReader:
    """Extract workbook content while preserving sheets, rows, and cell coordinates."""

    _OPENPYXL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}

    def read(self, source_path: Path) -> SourceDocument:
        path = source_path.expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)

        suffix = path.suffix.lower()
        if suffix in self._OPENPYXL_SUFFIXES:
            text = self._read_openpyxl_workbook(path)
        elif suffix == ".xls":
            text = self._read_xlrd_workbook(path)
        else:
            raise ValueError(f"Expected an Excel file, got: {path}")

        return SourceDocument.from_text(path=path, text=text, source_type="excel")

    def _read_openpyxl_workbook(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Reading .xlsx/.xlsm files requires openpyxl. Install project dependencies with "
                'pip install -e ".[dev]".'
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sections = [f"--- source type: excel workbook ---\n--- file: {path.name} ---"]
        for sheet in workbook.worksheets:
            sections.append(f"\n\n--- sheet: {sheet.title} ---")
            for row in sheet.iter_rows():
                cells = [
                    f"{cell.coordinate}: {_format_cell_value(cell.value)}"
                    for cell in row
                    if cell.value is not None and str(cell.value).strip() != ""
                ]
                if cells:
                    sections.append(" | ".join(cells))
        workbook.close()
        return "\n".join(sections).strip()

    def _read_xlrd_workbook(self, path: Path) -> str:
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError(
                "Reading .xls files requires xlrd. Install project dependencies with "
                'pip install -e ".[dev]".'
            ) from exc

        workbook = xlrd.open_workbook(path)
        sections = [f"--- source type: excel workbook ---\n--- file: {path.name} ---"]
        for sheet in workbook.sheets():
            sections.append(f"\n\n--- sheet: {sheet.name} ---")
            for row_index in range(sheet.nrows):
                cells = []
                for column_index in range(sheet.ncols):
                    value = sheet.cell_value(row_index, column_index)
                    if value is None or str(value).strip() == "":
                        continue
                    coordinate = f"R{row_index + 1}C{column_index + 1}"
                    cells.append(f"{coordinate}: {_format_cell_value(value)}")
                if cells:
                    sections.append(" | ".join(cells))
        return "\n".join(sections).strip()


def _format_cell_value(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
